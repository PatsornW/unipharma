from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

ACCENT = "1177CC"
HDR_FONT = Font(name="Tahoma", bold=True, color="FFFFFF", size=11)
HDR_FILL = PatternFill("solid", fgColor=ACCENT)
CELL_FONT = Font(name="Tahoma", size=10)
INPUT_FONT = Font(name="Tahoma", size=10, color="0000FF")
NOTE_FONT = Font(name="Tahoma", size=10, color="555555")
TITLE_FONT = Font(name="Tahoma", bold=True, size=13, color=ACCENT)
thin = Side(style="thin", color="D0D7DE")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

CATS = [
    ("CAT01","ยาปฏิชีวนะ","Antibiotics"),("CAT02","ยาแก้ปวดลดไข้","Analgesics"),
    ("CAT03","ยาระบบหัวใจ-หลอดเลือด","Cardiovascular"),("CAT04","ยาเบาหวาน","Antidiabetics"),
    ("CAT05","ยาระบบทางเดินหายใจ","Respiratory"),("CAT06","ยาระบบทางเดินอาหาร","Gastrointestinal"),
    ("CAT07","วิตามินและอาหารเสริม","Vitamins & Supplements"),("CAT08","ยาโรคผิวหนัง","Dermatology"),
    ("CAT09","ยาตา/หู/จมูก","Ophthalmology/ENT"),("CAT10","อุปกรณ์การแพทย์","Medical Devices"),
]

def style_header(ws, headers):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HDR_FONT; cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

def write_rows(ws, rows, start=2):
    for r, row in enumerate(rows, start):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = INPUT_FONT; cell.border = BORDER

# ---------------- DRUGS ----------------
wb = Workbook()
ws = wb.active; ws.title = "ยา (Drugs)"
d_headers = ["code","nameTH","nameEN","unit","catId","hasVat","costEx","sellEx",
             "stockPTN","stockRAM","stockCNX","minStock","supplierId"]
style_header(ws, d_headers)
write_rows(ws, [
    ["AMX001","อะม็อกซิซิลลิน 500มก.","Amoxicillin 500mg","เม็ด","ยาปฏิชีวนะ",0,18,38,2400,1800,1200,500,"SUP001"],
    ["PAR001","พาราเซตามอล 500มก.","Paracetamol 500mg","เม็ด","ยาแก้ปวดลดไข้",0,1.5,5,8000,6000,4000,2000,"SUP001"],
    ["VTC001","วิตามินซี 500มก.","Vitamin C 500mg","เม็ด","วิตามินและอาหารเสริม",1,4,12,3000,2000,1500,800,"SUP005"],
])
widths = [12,30,28,10,24,8,10,10,11,11,11,10,12]
for i, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = w

# data validations
vat_dv = DataValidation(type="list", formula1='"0,1"', allow_blank=True)
vat_dv.prompt = "0 = ไม่มี VAT, 1 = มี VAT 7%"; vat_dv.promptTitle = "hasVat"
ws.add_data_validation(vat_dv); vat_dv.add(f"F2:F1000")
cat_dv = DataValidation(type="list", formula1='"'+",".join(c[1] for c in CATS)+'"', allow_blank=True)
cat_dv.prompt = "เลือกหมวดหมู่ (หรือพิมพ์รหัส CAT01-CAT10)"; cat_dv.promptTitle = "catId"
ws.add_data_validation(cat_dv); cat_dv.add("E2:E1000")

# instructions sheet for drugs
info = wb.create_sheet("คำแนะนำ")
info["A1"] = "วิธีกรอกฐานข้อมูลยา"; info["A1"].font = TITLE_FONT
guide = [
    ("", ""),
    ("คอลัมน์", "ความหมาย / ตัวอย่าง"),
    ("code", "รหัสสินค้า (ห้ามซ้ำ ใช้เป็นกุญแจ) เช่น AMX001  [จำเป็น]"),
    ("nameTH", "ชื่อยาไทย  [จำเป็น]"),
    ("nameEN", "ชื่อยาอังกฤษ"),
    ("unit", "หน่วย เช่น เม็ด/ขวด/แคปซูล  [จำเป็น]"),
    ("catId", "หมวดหมู่ — ใส่ชื่อหรือรหัสก็ได้ (ดูตารางด้านล่าง)  [จำเป็น]"),
    ("hasVat", "0 = ไม่มี VAT, 1 = มี VAT 7%  [จำเป็น]"),
    ("costEx", "ต้นทุน (ไม่รวม VAT)  [จำเป็น]"),
    ("sellEx", "ราคาขาย (ไม่รวม VAT)  [จำเป็น]"),
    ("stockPTN", "สต็อกสาขาประตูน้ำ"),
    ("stockRAM", "สต็อกสาขารามคำแหง"),
    ("stockCNX", "สต็อกสาขาเชียงใหม่"),
    ("minStock", "สต็อกขั้นต่ำ (แจ้งเตือนใกล้หมด)"),
    ("supplierId", "รหัสผู้จัดจำหน่ายหลัก (ต้องตรงกับไฟล์ผู้จัดจำหน่าย)"),
    ("", ""),
    ("หมายเหตุ", "กำไร/มาร์จิ้น และราคารวม VAT ระบบคำนวณให้อัตโนมัติ ไม่ต้องกรอก"),
    ("", ""),
    ("รหัสหมวดหมู่ที่ระบบรู้จัก", ""),
    ("รหัส", "ชื่อหมวดหมู่"),
]
r = 1
for a, b in guide:
    r += 1
    info.cell(row=r, column=1, value=a).font = Font(name="Tahoma", bold=(a in ("คอลัมน์","หมายเหตุ","รหัส","รหัสหมวดหมู่ที่ระบบรู้จัก")), size=10)
    info.cell(row=r, column=2, value=b).font = NOTE_FONT
for code, th, en in CATS:
    r += 1
    info.cell(row=r, column=1, value=code).font = CELL_FONT
    info.cell(row=r, column=2, value=f"{th}  /  {en}").font = NOTE_FONT
info.column_dimensions["A"].width = 26; info.column_dimensions["B"].width = 60
wb.save("unipharma_drugs_template.xlsx")

# ---------------- SUPPLIERS ----------------
wb2 = Workbook()
ws2 = wb2.active; ws2.title = "ผู้จัดจำหน่าย (Suppliers)"
s_headers = ["id","name","nameEN","contact","phone","email","taxId","creditTerm","deliveryDays","rating","address"]
style_header(ws2, s_headers)
write_rows(ws2, [
    ["SUP001","บริษัท ยา-สยาม ฟาร์มาซี จำกัด","Yah-Siam Pharmacy Co. Ltd.","คุณสมชาย ใจดี","02-234-5678","order@yahsiam.co.th","0105562001001",30,2,4.7,"45/12 ถนนพระรามสี่ กรุงเทพ 10500"],
    ["SUP002","บริษัท ดีเคเอสเอช (ประเทศไทย) จำกัด","DKSH (Thailand) Ltd.","คุณวรรณา พานิช","02-345-6789","pharma@dksh.co.th","0105562002002",45,3,4.5,"1 อาคารเอ็มไพร์ทาวเวอร์ กรุงเทพ 10120"],
])
swidths = [12,32,28,18,16,24,16,11,13,8,40]
for i, w in enumerate(swidths, 1): ws2.column_dimensions[get_column_letter(i)].width = w

info2 = wb2.create_sheet("คำแนะนำ")
info2["A1"] = "วิธีกรอกผู้จัดจำหน่าย"; info2["A1"].font = TITLE_FONT
sguide = [
    ("", ""),
    ("คอลัมน์", "ความหมาย / ตัวอย่าง"),
    ("id", "รหัสผู้จัดจำหน่าย (ห้ามซ้ำ) เช่น SUP001  [จำเป็น]"),
    ("name", "ชื่อบริษัท (ไทย)  [จำเป็น]"),
    ("nameEN", "ชื่อบริษัท (อังกฤษ)"),
    ("contact", "ชื่อผู้ติดต่อ"),
    ("phone", "เบอร์โทร"),
    ("email", "อีเมล"),
    ("taxId", "เลขผู้เสียภาษี"),
    ("creditTerm", "เครดิต (วัน) เช่น 30"),
    ("deliveryDays", "ส่งภายใน (วัน) เช่น 2"),
    ("rating", "คะแนน 0-5 เช่น 4.7"),
    ("address", "ที่อยู่"),
    ("", ""),
    ("สำคัญ", "ใส่ผู้จัดจำหน่ายก่อน แล้วค่อยนำเข้ายา เพื่อให้ supplierId ตรงกัน"),
]
r = 1
for a, b in sguide:
    r += 1
    info2.cell(row=r, column=1, value=a).font = Font(name="Tahoma", bold=(a in ("คอลัมน์","สำคัญ")), size=10)
    info2.cell(row=r, column=2, value=b).font = NOTE_FONT
info2.column_dimensions["A"].width = 26; info2.column_dimensions["B"].width = 60
wb2.save("unipharma_suppliers_template.xlsx")
print("OK created both xlsx")
